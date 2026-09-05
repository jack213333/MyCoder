"""
系统测试执行器

从 JSON 文件加载系统测试用例，在沙箱中启动 MyCoder 执行测试，
通过 LLM 评判结果，并生成 Excel 报告。

参考：
- unit_test_runner.py: 整体结构、progress_callback、_case 属性、Excel 报告样式
- new_feature_runner.py: 沙箱执行、_build_actual_output、TestResult 数据模型
- regression_runner.py: _build_actual_output 逻辑（与 new_feature_runner 一致）
"""

from __future__ import annotations

import json
import logging
import time
from pathlib import Path
from typing import Callable, Optional

from ..models import TestCase, TestResult, TestStatus
from ..sandbox import SandboxManager
from ..judge import LLMJudge

logger = logging.getLogger(__name__)


class SystemTestRunner:
    """系统测试用例执行器

    从 JSON 文件加载系统测试用例，在沙箱中启动 MyCoder，
    发送 user_prompt 并收集输出，通过 LLM 评判 PASS / FAIL。
    """

    def __init__(self, sandbox_mgr: SandboxManager, judge: LLMJudge):
        self._sandbox_mgr = sandbox_mgr
        self._judge = judge

    # ------------------------------------------------------------------

    def execute(self,
                test_cases: list,
                myclaude_root: str | None = None,
                progress_callback: Optional[Callable] = None) -> list[TestResult]:
        """执行全部系统测试用例，返回结果列表

        Args:
            test_cases: 测试用例列表（支持 TestCase 对象或 dict）
            myclaude_root: MyCoder 源码根目录
            progress_callback: 可选回调，签名为 callback(idx, total, results)
                idx: 当前已完成的用例序号（1-based）
                total: 总用例数
                results: 已完成用例的结果列表

        Returns:
            TestResult 列表
        """
        results: list[TestResult] = []
        total = len(test_cases)

        for i, raw_case in enumerate(test_cases):
            # 归一化：dict → TestCase，统一用属性访问
            case = self._normalize_case(raw_case)
            logger.info("Running system-test case [id=%s] %s", case.id, case.description)
            result = self._run_one(case, myclaude_root)
            # 注入原始用例数据，供 Excel 报告使用（Pydantic v2 需绕过 __setattr__）
            object.__setattr__(result, "_case", raw_case)
            results.append(result)
            logger.info("Case [id=%s] -> %s", case.id, result.status)
            logger.info("\n\n")

            if progress_callback:
                progress_callback(i + 1, total, results)

        return results

    # ------------------------------------------------------------------

    @staticmethod
    def load_test_cases(json_path: str | Path) -> list[dict]:
        """从 JSON 文件加载系统测试用例。

        Args:
            json_path: JSON 文件路径

        Returns:
            测试用例 dict 列表
        """
        path = Path(json_path)
        if not path.exists():
            raise FileNotFoundError(f"测试用例文件不存在: {path}")

        with open(path, "r", encoding="utf-8") as f:
            cases = json.load(f)

        if not isinstance(cases, list):
            raise ValueError(f"测试用例文件格式错误，期望 list，得到 {type(cases)}")

        logger.info("Loaded %d system test cases from %s", len(cases), path)
        return cases

    # ------------------------------------------------------------------

    @staticmethod
    def _normalize_case(raw_case) -> TestCase:
        """将 dict 或 TestCase 对象归一化为 TestCase。

        如果 raw_case 是 dict，尝试直接转换为 TestCase。
        如果因额外字段（如 scenario_id、section）导致转换失败，
        则过滤掉不支持的字段后重试。
        """
        if isinstance(raw_case, TestCase):
            return raw_case

        if isinstance(raw_case, dict):
            try:
                return TestCase(**raw_case)
            except Exception:
                # 过滤掉 TestCase 可能不支持的字段
                known_fields = {
                    "id", "description", "user_prompt",
                    "expected_behavior", "check_type",
                }
                filtered = {k: v for k, v in raw_case.items() if k in known_fields}
                return TestCase(**filtered)

        raise TypeError(f"不支持的用例类型: {type(raw_case)}")

    # ------------------------------------------------------------------

    def _run_one(self, case: TestCase, myclaude_root: str | None) -> TestResult:
        """执行单个系统测试用例。

        在沙箱中启动 MyCoder，发送 user_prompt，
        收集输出并通过 LLM 评判结果。
        """
        t0 = time.perf_counter()
        sandbox = None

        try:
            # 1. 在沙箱中启动 MyCoder 并发送指令，获取结构化测试结果
            sandbox = self._sandbox_mgr.acquire(myclaude_root)
            std_out, std_err, exit_code, test_data = \
                sandbox.run_myclaude_command_with_test_output(
                    user_prompt=case.user_prompt,
                    myclaude_root=myclaude_root,
                )

            # 2. 构建评判 actual_output（优先使用结构化 JSON 的关键输出片段）
            actual_output = self._build_actual_output(std_out, test_data, std_err, exit_code)

            # 3. 确定 check_type
            check_type = getattr(case, "check_type", None) or "general"

            # 4. 调用评判 LLM
            judge_reason = ""
            if exit_code == 0:
                verdict_result = self._judge.evaluate(
                    expected=case.expected_behavior,
                    actual_output=actual_output,
                    context=case.description,
                    check_type=check_type,
                )
                judge_reason = verdict_result.get("reason", "")

                judge_verdict = verdict_result.get("verdict")
                if judge_verdict and judge_verdict in (
                    TestStatus.PASS, TestStatus.FAIL,
                    TestStatus.INCONCLUSIVE, TestStatus.ERROR,
                ):
                    verdict = judge_verdict
                elif verdict_result.get("pass"):
                    verdict = TestStatus.PASS
                else:
                    verdict = TestStatus.FAIL
            else:
                verdict = TestStatus.FAIL
                judge_reason = f"子进程退出码非零: {exit_code}"

            elapsed = round(time.perf_counter() - t0, 2)
            return TestResult(
                test_id=case.id,
                description=case.description,
                status=verdict,
                actual_output=actual_output[:3000] if actual_output else "(empty)",
                exit_code=exit_code,
                duration_seconds=elapsed,
                judge_reason=judge_reason,
            )

        except Exception as exc:
            elapsed = round(time.perf_counter() - t0, 2)
            logger.exception("System-test case [id=%s] crashed", case.id)
            return TestResult(
                test_id=case.id,
                description=case.description,
                status=TestStatus.ERROR,
                actual_output=str(exc)[:3000],
                exit_code=-1,
                duration_seconds=elapsed,
                judge_reason=f"执行异常: {exc}",
            )

        finally:
            # 只在成功 acquire 后才 release，防止操作未设置的容器
            if sandbox is not None:
                self._sandbox_mgr.release(sandbox)

    # ------------------------------------------------------------------

    @staticmethod
    def _strip_ansi(text: str) -> str:
        """去除 Rich ANSI 转义码，返回纯文本。"""
        import re
        # 匹配 ANSI 转义序列（CSI + OSC + 其他常见模式）
        # CSI: \x1b[ + 参数(数字;?等) + 字母终结符  —— 覆盖颜色、光标移动、清屏等
        # OSC: \x1b] + 内容 + \x07(BEL) 或 \x1b\\(ST)
        # 其他: \x1b( 或 \x1b) + 字符（字符集切换），\x1b= \x1b>（键盘模式）
        ansi_re = re.compile(
            r'\x1b\[[0-9;?]*[a-zA-Z]'    # CSI 序列（含 ?25l/h 等光标命令）
            r'|\x1b\][^\x07\x1b]*(?:\x07|\x1b\\)'  # OSC 序列
            r'|\x1b[()][AB012]'           # 字符集切换
            r'|\x1b[=>]'                  # 键盘模式
            r'|\x1b[78]'                  # 保存/恢复光标
        )
        cleaned = ansi_re.sub('', text)
        # 去除残留的控制字符（保留换行 \n 和制表符 \t）
        cleaned = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', cleaned)
        # 处理 \r：将 \r\n 替换为 \n，然后去除单独的 \r（Rich Live 刷新产生）
        cleaned = cleaned.replace('\r\n', '\n').replace('\r', '\n')
        # 合并连续空行（最多保留2行）
        cleaned = re.sub(r'\n{3,}', '\n\n', cleaned)
        return cleaned.strip()

    @staticmethod
    def _build_actual_output(std_out: str, test_data: dict | None, std_err: str = "",
                             exit_code: int = -1) -> str:
        """将结构化测试数据合并为评判 LLM 的输入。

        优先提取 LLM 的实际回答（key_outputs 和 assistant 回复），
        而非完整对话历史（含系统提示词和项目上下文），避免无关内容干扰评判。
        如果 JSON 不可用或内容为空，回退到清理 ANSI 码后的 stdout。
        """
        parts = []

        # 尝试从结构化 JSON 提取有效内容
        if test_data:
            # 0. 用户原始输入（仅指令本身，不含注入的项目上下文）
            user_input = test_data.get("user_original_input", "")
            if user_input:
                parts.append(f"【用户输入】\n{user_input}")

            # 1. LLM 输出的纯文本片段（最核心的评判依据）
            #    key_outputs 与 conversation_history 中的 assistant 回复内容重复，
            #    仅保留 key_outputs 即可，避免冗余信息干扰评判 LLM
            key_outputs = test_data.get("key_outputs", [])
            if key_outputs:
                parts.append("【LLM 输出】")
                for ko in key_outputs:
                    parts.append(ko[:1500])
            else:
                # key_outputs 为空时，从对话历史中提取 assistant 回复作为回退
                conversation_history = test_data.get("conversation_history", [])
                assistant_msgs = [
                    msg for msg in conversation_history
                    if msg.get("role") == "assistant" and msg.get("content", "").strip()
                ]
                if assistant_msgs:
                    parts.append("【LLM 输出】")
                    for msg in assistant_msgs:
                        content = msg.get("content", "")
                        parts.append(content[:1500])

            # 2. 工具调用摘要
            tool_calls = test_data.get("tool_calls", [])
            if tool_calls:
                parts.append("【工具调用序列】")
                for i, tc in enumerate(tool_calls):
                    parts.append(
                        f"[{i+1}] {tc.get('tool', '?')}: "
                        f"params={tc.get('params', {})}, "
                        f"result={tc.get('result', '')[:500]}"
                    )

            # 3. 系统消息（含 done 消息、自动结束等关键信息）
            info_messages = test_data.get("info_messages", [])
            if info_messages:
                parts.append("【系统消息】")
                for im in info_messages:
                    parts.append(im[:500])

            # 4. 错误信息
            error = test_data.get("error")
            if error:
                parts.append(f"【异常信息】\n{error}")

            # 5. 截断标记
            if test_data.get("is_truncated"):
                parts.append("【警告】\nLLM 输出被截断（max_tokens 不足）")

        # 如果结构化数据没有提取到任何有效内容，逐层回退
        if not parts:
            # 层1: test_data 中的 full_output（Rich Console 全部输出）
            full_output = test_data.get("full_output", "") if test_data else ""
            if full_output:
                cleaned = SystemTestRunner._strip_ansi(full_output)
                if cleaned:
                    parts.append("【MyCoder 终端输出(full)】")
                    parts.append(cleaned[:3500])

            # 层2: 原始 stdout（子进程标准输出）
            if not parts:
                cleaned_stdout = SystemTestRunner._strip_ansi(std_out)
                if cleaned_stdout:
                    parts.append("【MyCoder 终端输出(stdout)】")
                    parts.append(cleaned_stdout[:3500])

            # 层3: stderr（最后兜底）
            if not parts and std_err:
                cleaned_stderr = SystemTestRunner._strip_ansi(std_err)
                if cleaned_stderr:
                    parts.append("【stderr】")
                    parts.append(cleaned_stderr[:2000])

            # 层4: 所有源都为空，输出明确提示
            if not parts:
                parts.append("【警告】\nMyCoder 未产生任何可捕获的输出内容。"
                             "可能原因：子进程启动失败、Rich 输出被重定向、或异常导致提前退出。")

        # 附加退出码（始终输出，优先使用参数传入的 exit_code）
        parts.append(f"【退出码】\n{exit_code}")

        result = "\n\n".join(parts)[:4000]
        return result if result else "(empty)"

    # ------------------------------------------------------------------

    @staticmethod
    def generate_excel_report(results: list[TestResult],
                              myclaude_root: str | None = None,
                              output_dir: str | None = None) -> Path | None:
        """根据系统测试结果生成 Excel 报告，输出到指定目录。

        Args:
            results: TestResult 列表（每个元素需携带 _case 原始用例数据）
            myclaude_root: MyCoder 源码根目录（用于定位 config，output_dir 提供时优先使用）
            output_dir: 输出目录（优先使用；未提供时从 config 读取 logs_root）

        Returns:
            生成的 .xlsx 文件路径，失败时返回 None
        """
        import sys
        from datetime import datetime
        from pathlib import Path

        if output_dir:
            logs_root = Path(output_dir)
        else:
            # 尝试从 global_cfg 读取，失败时回退到 myclaude_root/log
            try:
                root = myclaude_root or str(Path(__file__).resolve().parents[3])
                if root not in sys.path:
                    sys.path.insert(0, root)
                from src.utility.config_loader import global_cfg
                logs_root = Path(global_cfg.base_path.logs_root)
            except Exception as cfg_err:
                logger.warning("Failed to load global_cfg: %s, falling back", cfg_err)
                if myclaude_root:
                    logs_root = Path(myclaude_root) / "log"
                else:
                    logs_root = Path.cwd() / "log"

        logs_root.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"MyCoder_System_Test_Report_{timestamp}.xlsx"
        filepath = logs_root / filename

        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Side, PatternFill
        except ImportError:
            logger.error("openpyxl not installed, cannot generate Excel report")
            return None

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "System Test Report"

            # 设置默认行高为 15，使内容更易阅读
            ws.sheet_format.defaultRowHeight = 15

            # 表头：用例原始列 + 测试结果列
            headers = [
                "id", "scenario_id", "section", "description",
                "user_prompt", "expected_behavior",
                "status", "exit_code", "duration_seconds",
                "actual_output", "judge_reason",
            ]
            ws.append(headers)

            for result in results:
                case = getattr(result, "_case", None)
                if case is None:
                    row = [
                        result.test_id, "", "", result.description,
                        "", "",
                        result.status.value if hasattr(result.status, "value") else str(result.status),
                        result.exit_code, result.duration_seconds,
                        result.actual_output,
                        result.judge_reason,
                    ]
                else:
                    row = [
                        case.get("id", result.test_id),
                        case.get("scenario_id", ""),
                        case.get("section", ""),
                        case.get("description", result.description),
                        case.get("user_prompt", ""),
                        case.get("expected_behavior", ""),
                        result.status.value if hasattr(result.status, "value") else str(result.status),
                        result.exit_code,
                        result.duration_seconds,
                        result.actual_output,
                        result.judge_reason,
                    ]
                # 强制所有单元格为字符串类型，防止 Excel 将以 = 开头的内容当作公式
                safe_row = []
                for val in row:
                    if val is None:
                        safe_row.append("")
                    elif isinstance(val, str):
                        # 防止 Excel 公式注入：以 = + - @ 开头的内容会被 Excel 当作公式
                        if val and val[0] in ('=', '+', '-', '@'):
                            safe_row.append("'" + val)
                        else:
                            safe_row.append(val)
                    else:
                        safe_row.append(str(val))
                ws.append(safe_row)

            # --- 样式定义 ---
            yahei_font = openpyxl.styles.Font(name="微软雅黑", size=11)
            header_font = openpyxl.styles.Font(name="微软雅黑", size=11, bold=True)
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # --- 应用全局样式 ---
            for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row_cells:
                    cell.font = yahei_font
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.border = thin_border

            # --- 首行样式 ---
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = center_align
                cell.fill = header_fill

            # --- 居中列：id(1)、scenario_id(2)、section(3)、status(7)、exit_code(8)、duration(9) ---
            center_columns = [1, 2, 3, 7, 8, 9]
            for col_idx in center_columns:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.alignment = center_align

            # --- 冻结首行 ---
            ws.freeze_panes = "A2"

            # --- 自动调整列宽 ---
            for col_cells in ws.columns:
                max_length = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

            wb.save(filepath)
            logger.info("System test Excel report saved to %s", filepath)
            return filepath
        except Exception as build_err:
            logger.error("Failed to generate Excel report: %s", build_err, exc_info=True)
            return None


if __name__ == "__main__":
    import argparse
    import sys

    from src.utility.config_loader import global_cfg

    parser = argparse.ArgumentParser(
        description="SystemTestRunner CLI — 直接执行系统测试用例"
    )
    parser.add_argument(
        "--json",
        required=True,
        help="系统测试用例 JSON 文件路径",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="报告输出目录（默认使用 config 中 logs_root）",
    )
    parser.add_argument(
        "--myclaude-root",
        default=None,
        help="MyCoder 源码根目录（默认使用 config 中 project_root）",
    )
    args = parser.parse_args()

    # ── 配置日志 ──
    logs_root = Path(args.output) if args.output else Path(global_cfg.base_path.logs_root)
    logs_root.mkdir(parents=True, exist_ok=True)

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(
                logs_root / "system_test_runner.log",
                encoding="utf-8",
            ),
        ],
    )

    # ── 加载测试用例 ──
    json_path = Path(args.json)
    if not json_path.exists():
        logger.error("JSON 文件不存在: %s", json_path)
        sys.exit(1)

    st_test_cases = SystemTestRunner.load_test_cases(json_path)
    logger.info("从 %s 加载了 %d 条系统测试用例", json_path, len(st_test_cases))

    # ── 初始化组件 ──
    judge = LLMJudge()
    sandbox_mgr = SandboxManager()
    runner = SystemTestRunner(sandbox_mgr=sandbox_mgr, judge=judge)

    myclaude_root = args.myclaude_root or global_cfg.base_path.project_root

    # ── 执行测试 ──
    results = runner.execute(
        test_cases=st_test_cases,
        myclaude_root=myclaude_root,
    )

    # ── 生成 Excel 报告 ──
    report_path = SystemTestRunner.generate_excel_report(
        results, output_dir=str(logs_root)
    )

    # ── 打印总结 ──
    passed = sum(1 for r in results if r.status == TestStatus.PASS)
    total = len(results)
    print("\n" + "=" * 60)
    print("  系统测试完成")
    print(f"  通过: {passed}  失败: {total - passed}  合计: {total}")
    if total:
        print(f"  通过率: {passed / total * 100:.1f}%")
    else:
        print("  无测试用例")
    print(f"  Excel 报告: {report_path}")
    print("=" * 60)
