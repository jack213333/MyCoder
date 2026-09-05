"""
单元测试执行器

直接 import 被测模块、调用被测函数、捕获返回值/异常，再调用 LLMJudge 评判结果。
复用现有的 LLMJudge 评判逻辑，不做端到端 CLI 调用。
"""

from __future__ import annotations

import importlib
import json
import logging
import time
import traceback
from pathlib import Path

from src.A2A.test.models import UnitTestCase, UnitTestResult, TestStatus
from src.A2A.test.judge import LLMJudge

logger = logging.getLogger(__name__)


class UnitTestRunner:
    """单元测试用例执行器"""


    def __init__(self, judge: LLMJudge):
        self._judge = judge

    # ------------------------------------------------------------------

    def execute(self,
                test_cases: list,
                myclaude_root: str | None = None,
                progress_callback: callable | None = None) -> list[UnitTestResult]:
        """执行全部单元测试用例，返回结果列表

        Args:
            progress_callback: 可选回调，签名为 callback(idx, total, results)
                idx: 当前已完成的用例序号（1-based）
                total: 总用例数
                results: 已完成用例的结果列表
        """
        results: list[UnitTestResult] = []
        total = len(test_cases)

        for i, raw_case in enumerate(test_cases):
            # 归一化：dict → UnitTestCase，统一用属性访问
            case = self._normalize_case(raw_case)
            logger.info("Running unit-test case [id=%s] %s", case.id, case.description)
            result = self._run_one(case, myclaude_root)
            # 注入原始用例数据，供 Excel 报告使用（Pydantic v2 需绕过 __setattr__）
            object.__setattr__(result, "_case", case)
            results.append(result)
            logger.info("Case [id=%s] -> %s", case.id, result.status)
            logger.info("\n\n")

            if progress_callback:
                progress_callback(i + 1, total, results)

        return results

    # ------------------------------------------------------------------

    @staticmethod
    def load_test_cases(json_path: str | Path) -> list[dict]:
        """从 JSON 文件加载单元测试用例。

        Args:
            json_path: JSON 文件路径

        Returns:
            测试用例 dict 列表
        """
        path = Path(json_path)
        if not path.exists():
            raise FileNotFoundError(f"测试用例文件不存在: {path}")

        with open(path, "r", encoding="utf-8") as f:
            cases = json.load(f)

        if not isinstance(cases, list):
            raise ValueError(f"测试用例文件格式错误，期望 list，得到 {type(cases)}")

        logger.info("Loaded %d unit test cases from %s", len(cases), path)
        return cases

    # ------------------------------------------------------------------

    @staticmethod
    def _normalize_case(raw_case) -> UnitTestCase:
        """将 dict 或 UnitTestCase 对象归一化为 UnitTestCase。

        如果 raw_case 是 dict，尝试直接转换为 UnitTestCase。
        如果因额外字段导致转换失败，
        则过滤掉不支持的字段后重试。
        """
        if isinstance(raw_case, UnitTestCase):
            return raw_case

        if isinstance(raw_case, dict):
            try:
                return UnitTestCase(**raw_case)
            except Exception:
                # 过滤掉 UnitTestCase 可能不支持的字段
                known_fields = {
                    "id", "description", "target_module",
                    "target_function", "test_input",
                    "expected_behavior", "check_type",
                }
                filtered = {k: v for k, v in raw_case.items() if k in known_fields}
                return UnitTestCase(**filtered)

        raise TypeError(f"不支持的用例类型: {type(raw_case)}")

    # ------------------------------------------------------------------

    def _run_one(self, case: UnitTestCase, myclaude_root: str | None) -> UnitTestResult:
        t0 = time.perf_counter()

        try:
            # 1. 动态导入被测模块并调用函数
            actual_output = self._invoke_target(
                target_module=case.target_module,
                target_function=case.target_function,
                test_input=case.test_input,
                param_types=getattr(case, "param_types", {}) or {},
                myclaude_root=myclaude_root,
            )

            # 2. 调用评判 LLM

            verdict_result = self._judge.evaluate(
                expected=case.expected_behavior,
                actual_output=actual_output,
                context=case.description,
                check_type=case.check_type or "general",
            )

            verdict = verdict_result.get("verdict")
            if verdict and verdict in (TestStatus.PASS, TestStatus.FAIL, TestStatus.INCONCLUSIVE, TestStatus.ERROR):
                status = verdict
            else:
                status = TestStatus.PASS if verdict_result.get("pass") else TestStatus.FAIL
            elapsed = round(time.perf_counter() - t0, 2)
            reason = verdict_result.get("reason", "") or "（评判 LLM 未返回理由）"

            return UnitTestResult(
                test_id=case.id,
                description=case.description,
                status=status,
                actual_output=actual_output[:500],
                reason=reason,
                duration_seconds=elapsed,
            )

        except Exception as exc:
            elapsed = round(time.perf_counter() - t0, 2)
            logger.exception("Unit-test case [id=%s] crashed", case.id)
            return UnitTestResult(
                test_id=case.id,
                description=case.description,
                status=TestStatus.ERROR,
                actual_output=traceback.format_exc(),
                reason=str(exc),
                duration_seconds=elapsed,
            )

    # ------------------------------------------------------------------

    @staticmethod
    def generate_excel_report(results: list[UnitTestResult],
                              myclaude_root: str | None = None,
                              output_dir: str | None = None) -> Path | None:
        """根据测试结果生成 Excel 报告，输出到指定目录。

        Args:
            results: UnitTestResult 列表（每个元素需携带 _case 原始用例数据）
            myclaude_root: MyCoder 源码根目录（用于定位 config，output_dir 提供时优先使用）
            output_dir: 输出目录（优先使用；未提供时从 config 读取 logs_root）

        Returns:
            生成的 .xlsx 文件路径，失败时返回 None
        """
        import sys
        from datetime import datetime
        from pathlib import Path

        if output_dir:
            logs_root = Path(output_dir)
        else:
            # 尝试从 global_cfg 读取，失败时回退到 myclaude_root/log
            try:
                root = myclaude_root or str(Path(__file__).resolve().parents[4])
                if root not in sys.path:
                    sys.path.insert(0, root)
                from src.utility.config_loader import global_cfg
                logs_root = Path(global_cfg.base_path.logs_root)
            except Exception as cfg_err:
                logger.warning("Failed to load global_cfg: %s, falling back", cfg_err)
                if myclaude_root:
                    logs_root = Path(myclaude_root) / "log"
                else:
                    logs_root = Path.cwd() / "log"

        logs_root.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"MyCoder_Unit_Test_Report_{timestamp}.xlsx"
        filepath = logs_root / filename

        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Side, PatternFill
        except ImportError:
            logger.error("openpyxl not installed, cannot generate Excel report")
            return None

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Unit Test Report"

            # 设置默认行高为 15，使内容更易阅读
            ws.sheet_format.defaultRowHeight = 15

            # 表头：用例原始列 + 测试结果列
            headers = [
                "id", "description", "target_module", "target_function",
                "test_input", "expected_behavior",
                "status", "actual_output", "reason", "duration_seconds",
            ]
            ws.append(headers)

            for result in results:
                case = getattr(result, "_case", None)
                if case is None:
                    row = [
                        result.test_id, result.description, "", "", "", "",
                        result.status.value if hasattr(result.status, "value") else str(result.status),
                        result.actual_output, result.reason, result.duration_seconds,
                    ]
                else:
                    row = [
                        case.id,
                        case.description,
                        case.target_module,
                        case.target_function,
                        case.test_input,
                        case.expected_behavior,
                        result.status.value if hasattr(result.status, "value") else str(result.status),
                        result.actual_output,
                        result.reason,
                        result.duration_seconds,
                    ]
                ws.append(row)

            # --- 样式定义 ---
            yahei_font = openpyxl.styles.Font(name="微软雅黑", size=11)
            header_font = openpyxl.styles.Font(name="微软雅黑", size=11, bold=True)
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # --- 应用全局样式 ---
            for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row_cells:
                    cell.font = yahei_font
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.border = thin_border

            # --- 首行样式 ---
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = center_align
                cell.fill = header_fill

            # --- A列(1)、G列(7)、J列(10) 左右居中 ---
            center_columns = [1, 7, 10]
            for col_idx in center_columns:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.alignment = center_align

            # --- 冻结首行 ---
            ws.freeze_panes = "A2"

            # --- 自动调整列宽 ---
            for col_cells in ws.columns:
                max_length = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

            wb.save(filepath)
            logger.info("Excel report saved to %s", filepath)
            return filepath
        except Exception as build_err:
            logger.error("Failed to generate Excel report: %s", build_err, exc_info=True)
            return None

    # ------------------------------------------------------------------

    @staticmethod
    def _invoke_target(target_module: str,
                       target_function: str,
                       test_input: str,
                       param_types: dict | None = None,
                       myclaude_root: str | None = None) -> str:
        """动态导入被测模块并调用函数，返回 repr(result) 或异常字符串。

        Args:
            target_module: 如 'src.utility.file_tool'
            target_function: 如 'resolve_path'
            test_input: 测试说明字符串，复杂参数需 Runner 内部构造（见 _build_args）
            param_types: 参数名→类型提示的映射字典，用于类型强制转换
            myclaude_root: MyCoder 源码根目录

        Returns:
            格式化的实际输出字符串（含返回值或异常信息）
        """
        # 确保项目根在 sys.path 中
        import sys
        root = myclaude_root or str(Path(__file__).resolve().parents[4])
        if root not in sys.path:
            sys.path.insert(0, root)

        # 动态导入模块
        mod = importlib.import_module(target_module)
        func = getattr(mod, target_function)

        import io

        # 构造参数（根据 test_input 解析）
        args, kwargs = UnitTestRunner._build_args(
            test_input, target_function, param_types or {}, target_module
        )

        # 判断是否需要捕获 stdout/stderr
        _PRINT_FUNCTIONS = {
            "print_error", "print_info", "print_user_input",
            "print_header", "print_timestamp", "print_welcome",
            "print_banner", "print_blank", "print_tool_call",
            "print_tool_result", "print_unknown_cmd",
            "typewriter_print", "typewriter_then_markdown",
            "typewriter_then_collapse",
            "show_history", "show_token_count", "show_status",
            "print_dir_table", "clear_screen",
            "_show_reasoning_folded", "_show_reasoning_expanded",
            "expand_reasoning",
            "display_progress_bar", "print_separator",
            "contextmanager",
        }
        _cap_stdout = target_function in _PRINT_FUNCTIONS or target_function.startswith("print_")

        if _cap_stdout:
            # 捕获 stdout 和 stderr
            captured_stdout = io.StringIO()
            captured_stderr = io.StringIO()
            import sys
            old_stdout = sys.stdout
            old_stderr = sys.stderr
            sys.stdout = captured_stdout
            sys.stderr = captured_stderr
            exc_info = None
            try:
                result = func(*args, **kwargs)
            except Exception as _e:
                exc_info = _e
                result = None
            finally:
                sys.stdout = old_stdout
                sys.stderr = old_stderr
            out_str = captured_stdout.getvalue()
            err_str = captured_stderr.getvalue()
            parts = []
            if out_str:
                # 截断过长输出
                if len(out_str) > 500:
                    out_str = out_str[:500] + "...(截断)"
                parts.append(f"stdout: {repr(out_str)}")
            if err_str:
                if len(err_str) > 500:
                    err_str = err_str[:500] + "...(截断)"
                parts.append(f"stderr: {repr(err_str)}")
            if exc_info is not None:
                parts.append(f"抛出异常: {type(exc_info).__name__}: {str(exc_info)[:300]}")
            else:
                parts.append(f"返回值: {repr(result)}")
            return " | ".join(parts)
        else:
            try:
                result = func(*args, **kwargs)
                return f"返回值: {repr(result)}"
            except Exception as _e:
                return f"抛出异常: {type(_e).__name__}: {str(_e)[:300]}"

    # ------------------------------------------------------------------

    @staticmethod
    def _build_args(test_input: str,
                    target_function: str,
                    param_types: dict | None = None,
                    target_module: str = "") -> tuple[list, dict]:
        """根据 test_input 描述构造函数参数。

        优先尝试统一的键值对格式：'key1' : value1, 'key2' : value2
        键名对应被测函数的参数名。解析成功后全部以关键字参数形式传递。
        根据 param_types 进行类型强制转换（int/float/bool/Path/NoneType/List/Dict）。
        如果无法解析为键值对，回退到函数特定的旧格式解析逻辑。
        
        target_module: 被测模块路径，用于解析无引号标识符为实际对象
        """
        import re

        if param_types is None:
            param_types = {}

        # ── 内部辅助函数在文件顶部定义 ──

        # execute_code_tool: test_input 描述工具调用，需构造复合 tool_dict
        if target_function == "execute_code_tool":
            tool_match = re.search(r'(file_view|create|str_replace|bash|done|use_skill)', test_input)
            tool_name = tool_match.group(1) if tool_match else "file_view"

            tool_params = {}

            path_match = re.search(r"path=['\"]([^'\"]+)['\"]", test_input)
            if path_match:
                tool_params["path"] = path_match.group(1)
            limit_match = re.search(r"limit=['\"]?(\d+)", test_input)
            if limit_match:
                tool_params["limit"] = int(limit_match.group(1))
            offset_match = re.search(r"offset=['\"]?(\d+)", test_input)
            if offset_match:
                tool_params["offset"] = int(offset_match.group(1))
            summary_match = re.search(r"summary=['\"]([^'\"]*)['\"]", test_input)
            if summary_match:
                tool_params["summary"] = summary_match.group(1)

            if tool_name == "create":
                body_match = re.search(
                    r"(?:body|content)=['\"]([^'\"]+)['\"]", test_input
                )
                tool_params["content"] = body_match.group(1) if body_match else ""

            if tool_name == "str_replace":
                old_match = re.search(r"old=['\"]([^'\"]*)['\"]", test_input)
                new_match = re.search(r"new=['\"]([^'\"]*)['\"]", test_input)
                if old_match:
                    tool_params["old"] = old_match.group(1)
                if new_match:
                    tool_params["new"] = new_match.group(1)

            if tool_name == "use_skill":
                name_match = re.search(r"name=['\"]([^'\"]+)['\"]", test_input)
                if name_match:
                    tool_params["name"] = name_match.group(1)

            tool_dict = {"llm_tool": tool_name, "params": tool_params}
            return [tool_dict], {}

        # append_tool_exec_result: 参数是嵌套列表/字典，不适合键值对
        if target_function == "append_tool_exec_result":
            return [], {"api_messages": [
                {"role": "system", "content": "初始系统提示词"},
                {"role": "user", "content": "用户输入"},
            ], "tool_exec_result": {"role": "user", "content": "工具执行结果"}}

        # resolve_path: test_input 形如 "绝对路径 'D:/...' 和相对路径 'test.py'"
        if target_function == "resolve_path":
            paths = re.findall(r"['\"]([^'\"]+)['\"]", test_input)
            if paths:
                return paths, {}
            # 回退到 KV 解析
            kwargs = UnitTestRunner._parse_test_input_kv(
                test_input, param_types
            )
            if kwargs:
                return [], kwargs
            return [test_input], {}

        # parse_tools: test_input 是主 content
        if target_function == "parse_tools":
            # 如果 test_input 是 KV 格式，检查是否有 'content' 键
            kwargs = UnitTestRunner._parse_test_input_kv(
                test_input, param_types
            )
            if kwargs:
                # 可能只有一个 content 参数，尝试展开
                keys = list(kwargs.keys())
                if keys == ["content"]:
                    return [kwargs["content"]], {}
                return [], kwargs
            return [test_input], {}

        # file_create: 旧自然语言格式回退
        if target_function == "file_create":
            paths = re.findall(r"([A-Za-z]:/\S+\.py)", test_input)
            if len(paths) >= 2:
                return [paths[0], "print('v2')"], {}
            # 尝试 KV 解析
            kwargs = UnitTestRunner._parse_test_input_kv(
                test_input, param_types
            )
            if kwargs:
                return [], kwargs
            return [test_input], {}

        # strip_thinking: test_input 是待处理的文本
        if target_function == "strip_thinking":
            return [test_input], {}

        # ── 通用键值对解析：'key1' : value1, 'key2' : value2 ──
        kwargs = UnitTestRunner._parse_test_input_kv(
            test_input, param_types, target_module
        )
        if kwargs:
            return [], kwargs

        # ── 最终回退：test_input 作为单参数字符串传入 ──
        return [test_input], {}

    # ------------------------------------------------------------------

    @staticmethod
    def _parse_test_input_kv(test_input: str,
                             param_types: dict | None = None,
                             target_module: str = "") -> dict | None:
        """通用键值对解析器：'key1' : value1, 'key2' : value2

        支持多种 value 格式：
        - 引号字符串: 'hello', "world", 'it\\'s'
        - 无引号字面量: None, True, False, 123, -0.5, 3.14
        - 无引号标识符: generator_func, re_escape_default
        - 简单列表: [1, 2, 3], [{'a': 1}]
        - 简单字典: {'key': 'value'}
        - 表达式: 'A' * 10000
        - 混合: 各参数值格式可以不同

        Args:
            test_input: 原始 test_input 字符串
            param_types: 参数类型提示字典 {param_name: type_hint}
            target_module: 被测模块路径，用于解析无引号标识符为实际对象

        Returns:
            解析成功返回 kwargs 字典，失败返回 None
        """
        import re
        import ast

        if param_types is None:
            param_types = {}

        def _coerce_value(key: str, val_str: str):
            """根据 param_types 将字符串值转为对应 Python 类型。"""
            hint = param_types.get(key, "")
            hint_lower = hint.lower().strip()
            # 剥离 Optional/Union 包装
            hint_lower = re.sub(r"^optional\[(.*)\]$", r"\1", hint_lower, flags=re.IGNORECASE)
            hint_lower = re.sub(r"^union\[(.*)\]$", r"\1", hint_lower, flags=re.IGNORECASE)

            # 如果 hint 包含 None 且值为 None → 返回 None
            if "none" in hint_lower.split(","):
                if val_str.strip().lower() == "none":
                    return None

            base_types = [t.strip().strip("[]") for t in hint_lower.split(",")]

            for bt in base_types:
                bt_lower = bt.lower()
                if bt_lower in ("int", "integer"):
                    try:
                        return int(float(val_str))
                    except (ValueError, TypeError):
                        pass
                elif bt_lower in ("float", "number"):
                    try:
                        return float(val_str)
                    except (ValueError, TypeError):
                        pass
                elif bt_lower in ("bool", "boolean"):
                    v_lower = val_str.strip().lower()
                    if v_lower in ("true", "1"):
                        return True
                    elif v_lower in ("false", "0"):
                        return False
                elif bt_lower in ("str", "string", "anystr"):
                    return val_str
                elif bt_lower in ("path", "pathlike", "purepath"):
                    from pathlib import Path
                    return Path(val_str) if val_str else None
            return val_str

        def _parse_value_string(raw_val: str) -> str:
            """解析引号包裹的字符串值，还原转义字符。
            
            处理单引号包围和双引号包围两种格式。
            如 "'hello'" -> "hello", '"world"' -> "world"
            """
            v = raw_val.strip()
            if len(v) >= 2 and v[0] == v[-1] and v[0] in ("'", '"'):
                inner = v[1:-1]
                # 还原常见转义
                inner = inner.replace("\\n", "\n").replace("\\t", "\t")
                inner = inner.replace("\\'", "'").replace('\\"', '"')
                return inner
            return v

        def _parse_value_any(raw_val: str, key: str = "") -> object:
            """解析单个 value 字符串为 Python 对象。
            
            支持：
            - None → Python None
            - True/False → Python bool
            - 数字 → int 或 float
            - 引号字符串 → str（去除外层引号）
            - 无引号标识符 → 尝试从 target_module 解析为实际对象，失败则保留字符串
            - [...] → 尝试 ast.literal_eval 解析
            - {...} → 尝试 ast.literal_eval 解析
            """
            v = raw_val.strip()
            if not v:
                return v

            # None
            if v.lower() == "none":
                return None

            # 布尔值
            if v.lower() == "true":
                return True
            if v.lower() == "false":
                return False

            # 数字
            try:
                # 优先尝试 int
                if re.match(r'^-?\d+$', v):
                    return int(v)
            except ValueError:
                pass
            try:
                return float(v)
            except ValueError:
                pass

            # 引号字符串
            if len(v) >= 2 and v[0] == v[-1] and v[0] in ("'", '"'):
                return _parse_value_string(v)

            # 列表 [...] 或 字典 {...}
            if (v.startswith("[") and v.endswith("]")) or \
               (v.startswith("{") and v.endswith("}")):
                try:
                    return ast.literal_eval(v)
                except (ValueError, SyntaxError):
                    pass

            # 表达式（如 'A' * 10000）或标识符（如 generator_func, regular_func）
            # 对纯标识符（不含运算符），尝试从 target_module 解析为实际对象
            if re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', v) and target_module:
                try:
                    mod = importlib.import_module(target_module)
                    resolved = getattr(mod, v, None)
                    if resolved is not None:
                        return resolved
                except (ImportError, AttributeError, ValueError):
                    pass
            # 无法解析则保留原字符串
            return v

        if not test_input or not isinstance(test_input, str):
            return None

        ti_stripped = test_input.strip()
        if not ti_stripped:
            return None

        # ── 步骤1: 按分隔符 'key' : 切分键值对 ──
        # 匹配模式: 'key' : 或 "key" :
        # 后面跟任意值，以逗号分隔下一个键值对或到达字符串末尾
        ti = ti_stripped
        kwargs = {}

        while ti:
            ti = ti.lstrip()
            if not ti:
                break

            # 匹配键名: 'key' : 或 "key" :
            key_match = re.match(r"""['"]([^'"]+)['"]\s*:\s*""", ti)
            if not key_match:
                return None  # 格式不匹配

            key = key_match.group(1)
            ti = ti[key_match.end():]

            # ── 步骤2: 解析值 ──
            # 值以逗号 + 下一个 'key' : 为分界，或用完整个字符串
            next_kv_pos = -1
            pos = 0
            depth_brace = 0    # {}
            depth_bracket = 0  # []
            depth_paren = 0    # ()
            quote_char = None  # ' 或 "

            while pos < len(ti):
                ch = ti[pos]

                if quote_char is not None:
                    if ch == '\\':
                        pos += 1  # 跳过转义字符
                    elif ch == quote_char:
                        quote_char = None
                    pos += 1
                elif ch in ("'", '"'):
                    quote_char = ch
                    pos += 1
                elif ch == '{':
                    depth_brace += 1
                    pos += 1
                elif ch == '}':
                    depth_brace = max(0, depth_brace - 1)
                    pos += 1
                elif ch == '[':
                    depth_bracket += 1
                    pos += 1
                elif ch == ']':
                    depth_bracket = max(0, depth_bracket - 1)
                    pos += 1
                elif ch == '(':
                    depth_paren += 1
                    pos += 1
                elif ch == ')':
                    depth_paren = max(0, depth_paren - 1)
                    pos += 1
                elif ch == ',' and depth_brace == 0 and depth_bracket == 0 and depth_paren == 0:
                    # 检查逗号后面是否是新的键值对
                    peek_pos = pos + 1
                    while peek_pos < len(ti) and ti[peek_pos].isspace():
                        peek_pos += 1
                    if peek_pos < len(ti) and ti[peek_pos] in ("'", '"'):
                        # 尝试匹配下一个键名
                        peek_rest = ti[peek_pos:]
                        if re.match(r"""['"][^'"]+['"]\s*:""", peek_rest):
                            # 确认是分隔符
                            next_kv_pos = pos
                            break
                    # 逗号是值的一部分
                    pos += 1
                else:
                    pos += 1

            if next_kv_pos >= 0:
                value_str = ti[:next_kv_pos].strip()
                ti = ti[next_kv_pos + 1:]  # 跳过逗号
            else:
                value_str = ti.strip()
                ti = ""

            # 解析值
            parsed_value = _parse_value_any(value_str)
            # 如果 value 是字符串且 param_types 中有该键，应用类型强制转换
            if isinstance(parsed_value, str):
                parsed_value = _coerce_value(key, parsed_value)
            kwargs[key] = parsed_value

        return kwargs if kwargs else None


if __name__ == "__main__":
    import argparse
    import sys

    from src.utility.config_loader import global_cfg

    parser = argparse.ArgumentParser(
        description="UnitTestRunner CLI — 直接执行单元测试用例"
    )
    parser.add_argument(
        "--json",
        required=True,
        help="单元测试用例 JSON 文件路径",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="报告输出目录（默认使用 config 中 logs_root）",
    )
    parser.add_argument(
        "--myclaude-root",
        default=None,
        help="MyCoder 源码根目录（默认使用 config 中 project_root）",
    )
    args = parser.parse_args()

    # ── 配置日志 ──
    logs_root = Path(args.output) if args.output else Path(global_cfg.base_path.logs_root)
    logs_root.mkdir(parents=True, exist_ok=True)

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(
                logs_root / "unit_test_runner.log",
                encoding="utf-8",
            ),
        ],
    )

    # ── 加载测试用例 ──
    json_path = Path(args.json)
    if not json_path.exists():
        logger.error("JSON 文件不存在: %s", json_path)
        sys.exit(1)

    ut_test_cases = UnitTestRunner.load_test_cases(json_path)
    logger.info("从 %s 加载了 %d 条单元测试用例", json_path, len(ut_test_cases))

    # ── 初始化组件 ──
    judge = LLMJudge()
    runner = UnitTestRunner(judge=judge)

    myclaude_root = args.myclaude_root or global_cfg.base_path.project_root

    # ── 执行测试 ──
    results = runner.execute(
        test_cases=ut_test_cases,
        myclaude_root=myclaude_root,
    )

    # ── 生成 Excel 报告 ──
    report_path = UnitTestRunner.generate_excel_report(
        results, output_dir=str(logs_root)
    )

    # ── 打印总结 ──
    passed = sum(1 for r in results if r.status == TestStatus.PASS)
    total = len(results)
    print("\n" + "=" * 60)
    print("  单元测试完成")
    print(f"  通过: {passed}  失败: {total - passed}  合计: {total}")
    if total:
        print(f"  通过率: {passed / total * 100:.1f}%")
    else:
        print("  无测试用例")
    print(f"  Excel 报告: {report_path}")
    print("=" * 60)
