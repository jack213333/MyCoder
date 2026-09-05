"""Token 统计模块。

在后台默默记录每次 LLM 调用的 token 消耗：
- 月度明细（XLSX，每月一个文件，每行对应一次 LLM 调用/turn）
- 汇总统计（1个XLSX文件，3个sheet：day/month/year）

Excel 公式说明：
  计费列使用 =E2*F2 格式的公式，单价列留空由用户手动填写。
  总数量列 = 命中缓存数量 + 未命中缓存数量 + 输出数量。
"""

import logging
import threading
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl

logger = logging.getLogger(__name__)

# 写文件锁（simple_chat 可能在记忆系统的线程中调用）
_write_lock = threading.Lock()

_STATS_DIR: Optional[Path] = None

# 明细表头：15 列（新增"总数量"列，位于"总计费"左侧）
_DETAIL_HEADERS = [
    "时间", "模型名称", "Query", "Turn",
    "输入(命中缓存)-数量", "输入(命中缓存)-单价", "输入(命中缓存)-计费",
    "输入(未命中缓存)-数量", "输入(未命中缓存)-单价", "输入(未命中缓存)-计费",
    "输出-数量", "输出-单价", "输出-计费",
    "总数量",
    "总计费",
]

# 汇总表头（3个sheet共用）
_SUMMARY_HEADERS = [
    "周期值", "模型名称",
    "输入(命中缓存)-数量", "输入(命中缓存)-单价", "输入(命中缓存)-计费",
    "输入(未命中缓存)-数量", "输入(未命中缓存)-单价", "输入(未命中缓存)-计费",
    "输出-数量", "输出-单价", "输出-计费",
    "总数量",
    "总计费",
]

# 汇总sheet名称
_SUMMARY_SHEETS = ["day", "month", "year"]


def _get_stats_dir() -> Path:
    """获取统计文件目录（单例缓存）。"""
    global _STATS_DIR
    if _STATS_DIR is not None:
        return _STATS_DIR
    from src.utility.config_loader import global_cfg
    _STATS_DIR = Path(global_cfg.base_path.project_root) / "token_statistics"
    _STATS_DIR.mkdir(parents=True, exist_ok=True)
    return _STATS_DIR


def get_stats_dir_path() -> str:
    """返回统计文件目录的绝对路径字符串，供 CLI 显示用。"""
    return str(_get_stats_dir())


def _get_monthly_detail_path() -> Path:
    """获取当月明细文件路径。"""
    return _get_stats_dir() / f"token_statistics_{datetime.now().strftime('%Y%m')}.xlsx"


def _get_summary_path() -> Path:
    """获取汇总统计文件路径（单个xlsx文件，含3个sheet）。"""
    return _get_stats_dir() / "token_statistics_summary.xlsx"


def _backup_old_file(filepath: Path) -> None:
    """备份旧格式文件。"""
    if not filepath.exists():
        return
    backup = filepath.with_name(filepath.stem + "_old" + filepath.suffix)
    filepath.rename(backup)
    logger.info(f"旧格式文件已备份: {backup}")


def _ensure_detail_workbook(filepath: Path) -> openpyxl.Workbook:
    """确保明细 xlsx 文件格式正确，返回可用的 Workbook。

    - 文件不存在 → 新建
    - 文件存在但表头不匹配 → 备份旧文件，新建
    - 文件存在且表头匹配 → 加载返回
    """
    # 备份同名旧CSV文件（如果存在）
    old_csv = filepath.with_suffix(".csv")
    if old_csv.exists():
        _backup_old_file(old_csv)

    if not filepath.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "明细"
        ws.append(_DETAIL_HEADERS)
        wb.save(filepath)
        return wb

    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        if header != _DETAIL_HEADERS:
            _backup_old_file(filepath)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "明细"
            ws.append(_DETAIL_HEADERS)
            wb.save(filepath)
        return wb
    except Exception:
        _backup_old_file(filepath)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "明细"
        ws.append(_DETAIL_HEADERS)
        wb.save(filepath)
        return wb


def _ensure_summary_workbook(filepath: Path) -> openpyxl.Workbook:
    """确保汇总 xlsx 文件格式正确，返回可用的 Workbook（含3个sheet）。

    - 文件不存在 → 新建，创建 day/month/year 三个sheet
    - 文件存在但格式不对 → 备份旧文件，新建
    - 文件存在且格式正确 → 加载返回
    """
    # 备份同名旧CSV文件（如果存在）
    for sheet_name in _SUMMARY_SHEETS:
        old_csv = _get_stats_dir() / f"token_statistics_summary_{sheet_name}.csv"
        if old_csv.exists():
            _backup_old_file(old_csv)

    if not filepath.exists():
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for sheet_name in _SUMMARY_SHEETS:
            ws = wb.create_sheet(sheet_name)
            ws.append(_SUMMARY_HEADERS)
        wb.save(filepath)
        return wb

    try:
        wb = openpyxl.load_workbook(filepath)
        need_rebuild = False
        for sheet_name in _SUMMARY_SHEETS:
            if sheet_name not in wb.sheetnames:
                need_rebuild = True
                break
            ws = wb[sheet_name]
            header = [cell.value for cell in ws[1]]
            if header != _SUMMARY_HEADERS:
                need_rebuild = True
                break
        if need_rebuild:
            _backup_old_file(filepath)
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            for sheet_name in _SUMMARY_SHEETS:
                ws = wb.create_sheet(sheet_name)
                ws.append(_SUMMARY_HEADERS)
            wb.save(filepath)
        return wb
    except Exception:
        _backup_old_file(filepath)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for sheet_name in _SUMMARY_SHEETS:
            ws = wb.create_sheet(sheet_name)
            ws.append(_SUMMARY_HEADERS)
        wb.save(filepath)
        return wb


def _update_summary_sheet(
    wb: openpyxl.Workbook, sheet_name: str,
    period_value: str, model_name: str,
    cached: int, uncached: int, output: int,
) -> None:
    """更新汇总统计 xlsx 中的一个 sheet。

    在内存中查找并累加已有行，找不到则追加新行。
    计费列和总数量列使用 Excel 公式，单价列留空由用户手动填写。
    """
    ws = wb[sheet_name]

    # 遍历已有行，查找匹配的 (周期值, 模型名称)
    for row_idx in range(2, ws.max_row + 1):
        existing_period = ws.cell(row=row_idx, column=1).value
        existing_model = ws.cell(row=row_idx, column=2).value
        if existing_period == period_value and existing_model == model_name:
            # 累加数量列（保留单价列不变）
            ws.cell(row=row_idx, column=3).value = int(ws.cell(row=row_idx, column=3).value or 0) + cached
            ws.cell(row=row_idx, column=6).value = int(ws.cell(row=row_idx, column=6).value or 0) + uncached
            ws.cell(row=row_idx, column=9).value = int(ws.cell(row=row_idx, column=9).value or 0) + output
            return

    # 未找到，追加新行
    new_row = ws.max_row + 1 if ws.max_row >= 1 else 2
    ws.cell(row=new_row, column=1).value = period_value
    ws.cell(row=new_row, column=2).value = model_name
    ws.cell(row=new_row, column=3).value = cached
    ws.cell(row=new_row, column=4).value = ""   # 单价留空
    ws.cell(row=new_row, column=5).value = f"=C{new_row}*D{new_row}"
    ws.cell(row=new_row, column=6).value = uncached
    ws.cell(row=new_row, column=7).value = ""
    ws.cell(row=new_row, column=8).value = f"=F{new_row}*G{new_row}"
    ws.cell(row=new_row, column=9).value = output
    ws.cell(row=new_row, column=10).value = ""
    ws.cell(row=new_row, column=11).value = f"=I{new_row}*J{new_row}"
    ws.cell(row=new_row, column=12).value = f"=C{new_row}+F{new_row}+I{new_row}"  # 总数量
    ws.cell(row=new_row, column=13).value = f"=E{new_row}+H{new_row}+K{new_row}"  # 总计费


def record_token_usage(
    model_name: str,
    prompt_tokens: int,
    cached_tokens: int,
    completion_tokens: int,
    query: str = "",
    turn: str = "",
) -> None:
    """记录一次 LLM 调用的 token 消耗。

    同时写入月度明细 xlsx（1行，含 Excel 公式）和汇总统计 xlsx（更新3个sheet）。
    线程安全，异常不抛出。

    Args:
        model_name: 模型名称
        prompt_tokens: 总输入 token 数
        cached_tokens: 缓存命中的 token 数
        completion_tokens: 输出 token 数
        query: 用户 Query 信息或 CLI 命令（如用户输入文本或 /mem extract）
        turn: 轮次标识，如 "turn1"、"turn2_followup1"、"CLI_COMMAND"
    """
    try:
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        uncached = max(0, prompt_tokens - cached_tokens)

        with _write_lock:
            # 1. 写入月度明细 xlsx
            detail_path = _get_monthly_detail_path()
            wb_detail = _ensure_detail_workbook(detail_path)
            ws = wb_detail.active
            new_row = ws.max_row + 1 if ws.max_row >= 1 else 2

            ws.cell(row=new_row, column=1).value = now_str
            ws.cell(row=new_row, column=2).value = model_name
            ws.cell(row=new_row, column=3).value = query
            ws.cell(row=new_row, column=4).value = turn
            ws.cell(row=new_row, column=5).value = cached_tokens
            ws.cell(row=new_row, column=6).value = ""
            ws.cell(row=new_row, column=7).value = f"=E{new_row}*F{new_row}"
            ws.cell(row=new_row, column=8).value = uncached
            ws.cell(row=new_row, column=9).value = ""
            ws.cell(row=new_row, column=10).value = f"=H{new_row}*I{new_row}"
            ws.cell(row=new_row, column=11).value = completion_tokens
            ws.cell(row=new_row, column=12).value = ""
            ws.cell(row=new_row, column=13).value = f"=K{new_row}*L{new_row}"
            ws.cell(row=new_row, column=14).value = f"=E{new_row}+H{new_row}+K{new_row}"  # 总数量
            ws.cell(row=new_row, column=15).value = f"=G{new_row}+J{new_row}+M{new_row}"  # 总计费
            wb_detail.save(detail_path)

            # 2. 更新汇总统计 xlsx（3个sheet）
            summary_path = _get_summary_path()
            wb_summary = _ensure_summary_workbook(summary_path)
            now = datetime.now()
            _update_summary_sheet(wb_summary, "day",
                now.strftime("%Y-%m-%d"), model_name, cached_tokens, uncached, completion_tokens)
            _update_summary_sheet(wb_summary, "month",
                now.strftime("%Y-%m"), model_name, cached_tokens, uncached, completion_tokens)
            _update_summary_sheet(wb_summary, "year",
                now.strftime("%Y"), model_name, cached_tokens, uncached, completion_tokens)
            wb_summary.save(summary_path)

    except Exception as e:
        logger.warning(f"记录 token 统计失败: {e}")


def get_token_summary() -> Dict:
    """读取汇总统计 xlsx 的 day sheet，返回当前累计的 token 统计。

    汇总 day sheet 中所有行的数据，得到全量累计值。

    Returns:
        dict: prompt_cache_hit, prompt_cache_miss, completion_tokens, total, stats_dir
    """
    summary_path = _get_summary_path()
    result = {
        "prompt_cache_hit": 0,
        "prompt_cache_miss": 0,
        "completion_tokens": 0,
        "total": 0,
        "stats_dir": str(_get_stats_dir()),
    }

    if not summary_path.exists():
        return result

    try:
        wb = openpyxl.load_workbook(summary_path, data_only=True)
        if "day" not in wb.sheetnames:
            return result
        ws = wb["day"]

        total_hit = 0
        total_miss = 0
        total_output = 0

        for row_idx in range(2, ws.max_row + 1):
            hit = ws.cell(row=row_idx, column=3).value
            miss = ws.cell(row=row_idx, column=6).value
            output = ws.cell(row=row_idx, column=9).value

            total_hit += int(hit or 0)
            total_miss += int(miss or 0)
            total_output += int(output or 0)

        result["prompt_cache_hit"] = total_hit
        result["prompt_cache_miss"] = total_miss
        result["completion_tokens"] = total_output
        result["total"] = total_hit + total_miss + total_output
    except Exception as e:
        logger.warning(f"读取 token 统计失败: {e}")

    return result
