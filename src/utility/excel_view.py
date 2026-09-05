"""
Excel 文件阅读工具，为 MyClaude 提供 <excel_view> 的支持能力。
"""
import sys
from pathlib import Path
from typing import Optional, Union

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore

# 兼容 xls 格式
try:
    import xlrd
except ImportError:
    xlrd = None  # type: ignore


def view_excel(
    file_path: Union[str, Path],
    sheet_name: str = "",
    start_row: Optional[int] = None,
    end_row: Optional[int] = None,
    start_col: Optional[int] = None,
    end_col: Optional[int] = None,
) -> str:
    """
    读取 Excel 文件（.xlsx / .xls）并返回可读的文本表格。

    Args:
        file_path: Excel 文件绝对路径
        sheet_name: 工作表名称或索引（默认第一个）
        start_row: 起始行（1-based），None 表示从第一行开始
        end_row: 结束行（1-based），None 表示到最后一行
        start_col: 起始列（1-based），None 表示从第一列开始
        end_col: 结束列（1-based），None 表示到最后一列

    Returns:
        格式化的表格字符串，包含行列标题；失败时返回错误信息。
    """
    file_path = Path(file_path)

    if not file_path.exists():
        return f"[错误] 文件不存在: {file_path}"

    suffix = file_path.suffix.lower()

    if suffix in (".xlsx", ".xlsm"):
        if openpyxl is None:
            return "[错误] 缺少 openpyxl 库，请执行 pip install openpyxl"
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = _pick_sheet(wb, sheet_name)
        if not ws:
            wb.close()
            return f"[错误] 工作簿中没有找到指定的工作表: {sheet_name}"

        rows = list(ws.iter_rows(min_row=1, values_only=False))
        result = _rows_to_text(rows, start_row, end_row, start_col, end_col)
        wb.close()
        return result

    elif suffix == ".xls":
        if xlrd is None:
            return "[错误] 缺少 xlrd 库，请执行 pip install xlrd"
        wb = xlrd.open_workbook(str(file_path))
        sh = _pick_sheet_xlrd(wb, sheet_name)
        if not sh:
            return f"[错误] 工作簿中没有找到指定的工作表: {sheet_name}"

        rows = []
        for r in range(sh.nrows):
            row_data = []
            for c in range(sh.ncols):
                row_data.append(sh.cell_value(r, c))
            rows.append(row_data)
        result = _raw_rows_to_text(rows, start_row, end_row, start_col, end_col)
        return result

    else:
        return f"[错误] 不支持的文件格式: {suffix}，仅支持 .xlsx / .xls"


def _pick_sheet(wb, sheet_name: str):
    """选择 openpyxl 工作表"""
    if not sheet_name:
        return wb.active
    try:
        # 尝试按索引
        idx = int(sheet_name)
        return wb.worksheets[idx]
    except (ValueError, IndexError):
        pass
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return None


def _pick_sheet_xlrd(wb, sheet_name: str):
    """选择 xlrd 工作表"""
    if not sheet_name:
        return wb.sheet_by_index(0)
    try:
        idx = int(sheet_name)
        return wb.sheet_by_index(idx)
    except (ValueError, IndexError):
        pass
    # xlrd 只能按索引，无法按名称直接获取；尝试遍历
    for name in wb.sheet_names():
        if name == sheet_name:
            return wb.sheet_by_name(name)
    return None


def _rows_to_text(rows: list, start_row, end_row, start_col, end_col) -> str:
    """将 openpyxl 行对象转换为文本表格"""
    data = []
    max_row_idx = len(rows)
    if max_row_idx == 0:
        return "[空表]"

    # 提取所有行的值和列数
    raw_rows = []
    max_col_cnt = 0
    for row in rows:
        vals = [cell.value for cell in row]
        raw_rows.append(vals)
        max_col_cnt = max(max_col_cnt, len(vals))

    return _raw_rows_to_text(raw_rows, start_row, end_row, start_col, end_col)


def _raw_rows_to_text(raw_rows: list, start_row, end_row, start_col, end_col) -> str:
    """通用行数据转文本表格"""
    if not raw_rows:
        return "[空表]"

    total_rows = len(raw_rows)
    # 确定实际行列范围
    srow = _clamp(start_row, 1, total_rows) - 1 if start_row else 0
    erow = _clamp(end_row, 1, total_rows) if end_row else total_rows

    if srow >= total_rows or srow >= erow:
        return f"[范围无效] 起始行 {start_row} 超出数据范围 (共 {total_rows} 行)"

    sliced = raw_rows[srow:erow]
    if not sliced:
        return "[空范围]"

    # 计算列数
    max_cols = max(len(r) for r in sliced)

    scol = _clamp(start_col, 1, max_cols) - 1 if start_col else 0
    ecol = _clamp(end_col, 1, max_cols) if end_col else max_cols

    # 列裁剪并补齐长度
    final_cols = []
    for r in sliced:
        # 扩展到 max_cols 长度，填充空字符串
        extended = list(r) + [""] * (max_cols - len(r))
        final_cols.append(extended[scol:ecol])

    # 生成字符串
    lines = []
    # 表头：行号 ｜ col1 ｜ col2 ...
    col_count = ecol - scol
    header = "行号 | " + " | ".join(f"C{scol + i + 1}" for i in range(col_count))
    lines.append(header)
    lines.append("-" * len(header))

    for idx, row_data in enumerate(final_cols):
        row_label = srow + idx + 1
        cleaned = [_format_cell(x) for x in row_data]
        line = f"{row_label:4d} | " + " | ".join(cleaned)
        lines.append(line)

    return "\n".join(lines)


def _format_cell(value) -> str:
    """格式化单元格为可读字符串"""
    if value is None:
        return ""
    if isinstance(value, float):
        # 避免 1.0 之类的显示过多小数点
        return f"{value:g}"
    return str(value)


def _clamp(val, lo, hi):
    """取值限制在 [lo, hi]"""
    if val is None:
        return None
    return max(lo, min(val, hi))


# 用于直接在控制台测试
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python excel_view.py <excel文件路径> [sheet名] [起始行] [结束行] [起始列] [结束列]")
        sys.exit(1)

    fpath = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else ""
    sr = int(sys.argv[3]) if len(sys.argv) > 3 else None
    er = int(sys.argv[4]) if len(sys.argv) > 4 else None
    sc = int(sys.argv[5]) if len(sys.argv) > 5 else None
    ec = int(sys.argv[6]) if len(sys.argv) > 6 else None

    output = view_excel(fpath, sheet, sr, er, sc, ec)
    print(output)
